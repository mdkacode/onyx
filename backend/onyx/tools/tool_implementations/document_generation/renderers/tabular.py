"""CSV and XLSX rendering.

Tabular formats consume the tables in a `DocumentSpec` and ignore its prose --
a five-section narrative report has no faithful CSV representation, so
pretending otherwise would produce a confusing file rather than a useful one.
A spec with no tables is rejected loudly for these formats.
"""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from typing_extensions import override

from onyx.tools.tool_implementations.document_generation.disclaimer import (
    DISCLAIMER_TEXT_LONG,
)
from onyx.tools.tool_implementations.document_generation.models import DocumentSpec
from onyx.tools.tool_implementations.document_generation.models import RenderedDocument
from onyx.tools.tool_implementations.document_generation.models import TableData
from onyx.tools.tool_implementations.document_generation.renderers.base import (
    DocumentRenderer,
)

# Excel and Sheets evaluate a cell as a formula when it starts with any of
# these, so an LLM-authored cell is a live injection vector -- the classic
# payload being =HYPERLINK(...) or a DDE call that runs on open.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _is_numeric(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def sanitize_cell(value: object) -> str:
    """Neutralize spreadsheet formula injection without mangling real data.

    Prefixing a single quote makes the spreadsheet treat the value as literal
    text. Negative numbers are deliberately exempt: `-5` starts with a trigger
    character but is data, and quoting it would break every numeric column
    that happens to contain one.
    """
    text = "" if value is None else str(value)
    if not text:
        return text
    if text.startswith(_FORMULA_TRIGGERS) and not _is_numeric(text):
        return "'" + text
    return text


class _TabularRenderer(DocumentRenderer):
    IS_TABULAR = True

    @staticmethod
    def _require_tables(spec: DocumentSpec) -> list[tuple[str, TableData]]:
        tables = spec.tables()
        if not tables:
            raise ValueError(
                "No table data to export. CSV and XLSX need at least one "
                "section containing a `table`; re-issue the call with the data "
                "as a table, or choose the 'pdf' or 'docx' format for prose."
            )
        return tables


class CsvRenderer(_TabularRenderer):
    FORMAT = "csv"
    MIME_TYPE = "text/csv"
    EXTENSION = "csv"

    @override
    def render(self, spec: DocumentSpec) -> RenderedDocument:
        tables = self._require_tables(spec)

        # CSV is a single-table format. Exporting only the first keeps the file
        # a valid CSV that every tool can parse; concatenating tables with
        # different column counts would not be. The dropped count is reported
        # back so the model can point the user at XLSX.
        heading, table = tables[0]
        notes: list[str] = []
        if len(tables) > 1:
            notes.append(
                f"Exported only the first of {len(tables)} tables "
                f"('{heading}') — CSV holds one table. Use 'xlsx' for all of them."
            )

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([sanitize_cell(h) for h in table.headers])

        width = len(table.headers)
        for row in table.rows:
            padded = [row[i] if i < len(row) else "" for i in range(width)]
            writer.writerow([sanitize_cell(c) for c in padded])

        # CSV has no footer or metadata channel, so the provenance notice goes
        # in a trailing comment row -- visible when opened, trivially dropped
        # by anything parsing the file programmatically.
        writer.writerow([])
        writer.writerow([f"# {DISCLAIMER_TEXT_LONG}"])

        return RenderedDocument(
            content=buffer.getvalue().encode("utf-8-sig"),
            mime_type=self.MIME_TYPE,
            extension=self.EXTENSION,
            unit_count=len(table.rows),
            unit_label="rows",
            notes=notes,
        )


class XlsxRenderer(_TabularRenderer):
    FORMAT = "xlsx"
    MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    EXTENSION = "xlsx"

    # Excel's hard limit on sheet-name length, and the characters it forbids.
    _MAX_SHEET_NAME = 31
    _ILLEGAL_SHEET_CHARS = str.maketrans({c: "-" for c in r"[]:*?/\\"})

    @classmethod
    def _sheet_name(cls, heading: str, index: int, taken: set[str]) -> str:
        base = heading.translate(cls._ILLEGAL_SHEET_CHARS).strip()
        base = base[: cls._MAX_SHEET_NAME] or f"Table {index + 1}"
        name = base
        suffix = 2
        while name.lower() in taken:
            trimmed = base[: cls._MAX_SHEET_NAME - len(str(suffix)) - 1]
            name = f"{trimmed} {suffix}"
            suffix += 1
        taken.add(name.lower())
        return name

    @override
    def render(self, spec: DocumentSpec) -> RenderedDocument:
        tables = self._require_tables(spec)

        workbook = Workbook()
        # Workbook() ships with one blank sheet; drop it so the first real
        # table does not land on a sheet named "Sheet".
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)  # ty: ignore[invalid-argument-type]
        taken: set[str] = set()

        for index, (heading, table) in enumerate(tables):
            sheet = workbook.create_sheet(self._sheet_name(heading, index, taken))
            sheet.append([sanitize_cell(h) for h in table.headers])
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            width = len(table.headers)
            for row in table.rows:
                padded = [row[i] if i < len(row) else "" for i in range(width)]
                sheet.append([sanitize_cell(c) for c in padded])

            sheet.append([])
            sheet.append([DISCLAIMER_TEXT_LONG])

        buffer = io.BytesIO()
        workbook.save(buffer)

        return RenderedDocument(
            content=buffer.getvalue(),
            mime_type=self.MIME_TYPE,
            extension=self.EXTENSION,
            unit_count=len(tables),
            unit_label="sheets",
        )
